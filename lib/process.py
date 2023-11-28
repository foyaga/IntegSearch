# -*- coding = utf-8 -*-
# @Time  : 2023/2/4 11:21
# @Author: Ifory
# @File  : process.py

import os

import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment
from prettytable import PrettyTable

from config import sqldb, database
from module.Finger import Finger
from module.Finger.config.data import logging


class Process(object):
    def __init__(self, result, keyword, searchType):
        self.result = result
        self.keyword = keyword
        self.searchType = searchType
        if not os.path.exists("result"):
            os.makedirs("result")

    def showScreen(self):
        table = PrettyTable(['HOST', 'IP', '端口', '协议', '国家', '域名', '标题'])
        table.align = 'l'
        for res in self.result:
            table.add_row([res[0], res[1], res[2], res[3], res[4], res[5], res[6]])
        print(table)

    def saveExcel(self, nowTime):
        file_path = f"result/{self.searchType}_{nowTime}.xlsx"

        TitleStyle = NamedStyle(name='TitleStyle',
                                font=Font(name='宋体', size=11, bold=True),
                                alignment=Alignment(horizontal='left', vertical='center'))

        BodyStyle = NamedStyle(name='BodyStyle',
                               alignment=Alignment(horizontal='left', vertical='center'))

        if not os.path.exists(file_path):
            keyword = self.keyword.replace(":", "_").replace("*", "").replace("\\", "").replace("/", "").replace("?", "").replace("[", "").replace("]", "")[:30]
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet(keyword, 0)

            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 10
            worksheet.column_dimensions['D'].width = 10
            worksheet.column_dimensions['E'].width = 25
            worksheet.column_dimensions['F'].width = 35
            worksheet.column_dimensions['G'].width = 60
            worksheet.column_dimensions['H'].width = 60

            col = ['URL', 'IP', '端口', '协议', '国家', '域名', '标题', '关键字']
            worksheet.append(col)
            for res in self.result:
                worksheet.append(res)

            # 标题样式
            for c in range(1, 9):
                worksheet.cell(1, c).style = TitleStyle
            worksheet.freeze_panes = 'A2'  # 冻结首行

            # 正文样式
            for r in range(2, len(self.result) + 2):
                worksheet.cell(r, 3).style = BodyStyle

            workbook.save(filename=f'result/{self.searchType}_{nowTime}.xlsx')
            logging.info(f'搜索结果文件输出路径为:{file_path}\n')
        else:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            max_row = worksheet.max_row  # 获得行数
            for res in self.result:
                worksheet.append(res)

            for r in range(max_row + 1, max_row + len(self.result) + 1):
                worksheet.cell(r, 3)

            workbook.save(file_path)
            logging.info(f'搜索结果文件输出路径为:{file_path}\n')

    def saveDatabase(self):
        database.CHECK_DB()
        db = sqldb.DB()
        db.query("use IntegSearch;")

        for res in self.result:
            sql = f"insert into {self.searchType} (host,ip,port,protocol,country,domain,title) values ('{res[0]}','{res[1]}','{res[2]}','{res[3]}','{res[4]}','{res[5]}','{res[6]}');"
            db.exec(sql)
        logging.info(f'搜索结果数据库保存成功！')

    def finger(self, output):
        Finger.main(self.result, output)
